"""
workbook_facts.py
------------------
Session 3, build plan 3.1: typed fact tables from the three pinned LNG
workbooks. Unlike ``workbook_reader.read_project_capacity`` (session 2, one
named capacity column only) this reads every column plan 3.1 names, trimming
on all-null rows and never trusting the sheet dimension, so a silent parse
failure surfaces as a wrong row count rather than a wrong downstream number.

Country resolution happens elsewhere (the applied alias crosswalk); these
readers keep the raw ``*_raw`` country strings alongside a synthetic surrogate
row id, because no natural business key survives across a project's trains or
a contract's Seller/Buyer/Start combination without risking a collision this
session has no evidence to rule out.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from lt_lng_flows.ingest.workbook_reader import (
    WORKBOOK_SPECS,
    _find_header_row,
    _last_nonblank_row,
)

LIQ_SPEC = WORKBOOK_SPECS[0]
REGAS_SPEC = WORKBOOK_SPECS[1]
CONTRACT_SPEC = WORKBOOK_SPECS[2]

# Column name -> fact table column name. Renamed to ascii lower snake case
# (CLAUDE.md) on read; the workbook's own header text is never used as a
# column identifier downstream.
_PROJECT_COLUMN_MAP_LIQ = {
    "Region": "region",
    "Country": "country_raw",
    "ISO 2-letter code": "iso2_raw",
    "Project": "project",
    "Trains": "trains",
    "Company": "company",
    "Start date": "start_date",
    "Start Year": "start_year",
    "MTPA": "mtpa",
    "bcf/d": "bcf_per_d",
    "bcma": "bcma",
    "Status": "status",
    "Type": "type",
    "FTA": "fta",
    "Non-FTA": "non_fta",
    "FERC Approval": "ferc_approval",
    "Export Licence": "export_licence",
    "FID Date": "fid_date",
}
_PROJECT_COLUMN_MAP_REGAS = {
    "Region": "region",
    "Country": "country_raw",
    "ISO 2-letter code": "iso2_raw",
    "Project": "project",
    "Trains": "trains",
    "Company": "company",
    "Start date": "start_date",
    "Start Year": "start_year",
    "MTPA": "mtpa",
    "bcf/d": "bcf_per_d",
    "bcma": "bcma",
    "Status": "status",
    "Type": "type",
}
_CONTRACT_COLUMN_MAP = {
    "Export country": "exporter_raw",
    "Supply Area": "supply_area",
    "Import country": "importer_raw",
    "Demand Area": "demand_area",
    "Loading Point": "loading_point",
    "Liquefaction project": "liquefaction_project",
    "Project partners": "project_partners",
    "Liquefaction project capacity": "liquefaction_project_capacity",
    "FID status": "fid_status",
    "Seller": "seller",
    "Buyer": "buyer",
    "bcm": "bcm",
    "MTPA": "mtpa",
    "Contract Start ": "contract_start",
    "Contract End": "contract_end",
    "Duration (years)": "duration_years",
    "Status": "status",
    "Delivery type": "delivery_type",
    "Destination flexibility": "destination_flexibility",
    "Agreement Type": "agreement_type",
    "Signed Date": "signed_date",
    "Pricing Index": "pricing_index",
    "Pricing slopes/ Price": "pricing_slope_or_price",
    "Contract price source": "contract_price_source",
    "Contract price confidence": "contract_price_confidence",
    "Notes": "notes",
}


def _read_full_rows(
    path: Path, sheet_name: str, required_columns: tuple[str, ...], column_map: dict[str, str]
) -> tuple[pd.DataFrame, int]:
    """Read every mapped column of ``sheet_name``, trimming on all-null rows.
    Returns (dataframe, header_row). Never trusts the sheet dimension
    (openpyxl / Excel report ~3500 rows on all three workbooks because of
    trailing formatting, per plan 3.1); the data range is bounded by the
    last row carrying any non-null value, exactly as ``workbook_reader``
    does for its narrower reads.
    """
    if not path.is_file():
        raise FileNotFoundError(f"workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: expected data sheet '{sheet_name}' not found; "
                f"sheets present are {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        header_row, header_values = _find_header_row(ws, required_columns)
        last_row = _last_nonblank_row(ws, header_row)
        col_index = {name: idx for idx, name in enumerate(header_values) if name is not None}

        missing = [c for c in column_map if c not in col_index]
        if missing:
            raise ValueError(
                f"{path.name} sheet '{sheet_name}': expected columns {missing} not found "
                f"in header row {header_row}"
            )

        records = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, values_only=True):
            if all(v is None for v in row):
                continue
            record = {out_name: row[col_index[in_name]] for in_name, out_name in column_map.items()}
            records.append(record)
        df = pd.DataFrame(records, columns=list(column_map.values()))
    finally:
        wb.close()
    return df, header_row


def read_fact_liq_project(path: Path) -> pd.DataFrame:
    """``fact_liq_project``: one row per liquefaction project/train, surrogate
    ``liq_project_row_id``, raw country strings (resolved by the caller
    through the applied crosswalk, never here)."""
    df, _ = _read_full_rows(
        path,
        LIQ_SPEC.data_sheet.sheet_name,
        LIQ_SPEC.data_sheet.required_columns,
        _PROJECT_COLUMN_MAP_LIQ,
    )
    df.insert(0, "liq_project_row_id", range(1, len(df) + 1))
    return df


def read_fact_regas_project(path: Path) -> pd.DataFrame:
    """``fact_regas_project``: one row per regas project/train, surrogate
    ``regas_project_row_id``."""
    df, _ = _read_full_rows(
        path,
        REGAS_SPEC.data_sheet.sheet_name,
        REGAS_SPEC.data_sheet.required_columns,
        _PROJECT_COLUMN_MAP_REGAS,
    )
    df.insert(0, "regas_project_row_id", range(1, len(df) + 1))
    return df


def read_fact_lng_contract(path: Path) -> pd.DataFrame:
    """``fact_lng_contract``: one row per contract, surrogate
    ``contract_row_id``."""
    df, header_row = _read_full_rows(
        path,
        CONTRACT_SPEC.data_sheet.sheet_name,
        CONTRACT_SPEC.data_sheet.required_columns,
        _CONTRACT_COLUMN_MAP,
    )
    df.insert(0, "contract_row_id", range(1, len(df) + 1))
    df.attrs["header_row"] = header_row
    return df


def assert_row_counts(
    fact_liq_project: pd.DataFrame,
    fact_regas_project: pd.DataFrame,
    fact_lng_contract: pd.DataFrame,
    constants: dict,
) -> None:
    """Build plan 3.1: row counts and country counts against the pinned
    202608 vintage. Raises and names the file, the sheet and both numbers on
    a mismatch, per the session 3 starting prompt.
    """
    cfg = constants["workbook_row_counts"]

    def _check(label, actual, expected):
        if actual != expected:
            raise AssertionError(
                f"3.1 row count assertion failed for {label}: expected {expected}, found {actual}"
            )

    _check(
        f"{cfg['liquefaction']['sheet']} row count",
        len(fact_liq_project),
        cfg["liquefaction"]["expected_rows"],
    )
    _check(
        f"{cfg['liquefaction']['sheet']} country count",
        fact_liq_project["country_raw"].nunique(),
        cfg["liquefaction"]["expected_countries"],
    )
    _check(
        f"{cfg['regas']['sheet']} row count",
        len(fact_regas_project),
        cfg["regas"]["expected_rows"],
    )
    _check(
        f"{cfg['regas']['sheet']} country count",
        fact_regas_project["country_raw"].nunique(),
        cfg["regas"]["expected_countries"],
    )
    _check(
        f"{cfg['contracts']['sheet']} row count",
        len(fact_lng_contract),
        cfg["contracts"]["expected_rows"],
    )
    header_row = fact_lng_contract.attrs.get("header_row")
    if header_row is not None:
        _check(
            f"{cfg['contracts']['sheet']} header row",
            header_row,
            cfg["contracts"]["header_row"],
        )


def assert_contract_distributions(fact_lng_contract: pd.DataFrame, constants: dict) -> None:
    """Build plan 3.1: contract distributions (status, delivery type, export
    and import country) against the pinned 202608 vintage."""
    cfg = constants["contract_distributions"]

    def _check_counts(label, series, expected: dict):
        actual = series.value_counts(dropna=False).to_dict()
        actual = {(k if pd.notna(k) else "-"): v for k, v in actual.items()}
        for key, expected_count in expected.items():
            found = actual.get(key, 0)
            if found != expected_count:
                raise AssertionError(
                    f"3.1 distribution assertion failed for {label}={key!r}: "
                    f"expected {expected_count}, found {found}"
                )

    _check_counts("status", fact_lng_contract["status"], cfg["status"])
    _check_counts("delivery_type", fact_lng_contract["delivery_type"], cfg["delivery_type"])

    export_cfg = cfg["export_country"]
    n_export = fact_lng_contract["exporter_raw"].nunique()
    if n_export != export_cfg["distinct_values"]:
        raise AssertionError(
            f"3.1 distribution assertion failed for export_country distinct values: "
            f"expected {export_cfg['distinct_values']}, found {n_export}"
        )
    n_portfolio = (fact_lng_contract["exporter_raw"] == "Portfolio").sum()
    if n_portfolio != export_cfg["Portfolio"]:
        raise AssertionError(
            f"3.1 distribution assertion failed for export_country=Portfolio: "
            f"expected {export_cfg['Portfolio']}, found {n_portfolio}"
        )

    import_cfg = cfg["import_country"]
    n_import = fact_lng_contract["importer_raw"].nunique()
    if n_import != import_cfg["distinct_values"]:
        raise AssertionError(
            f"3.1 distribution assertion failed for import_country distinct values: "
            f"expected {import_cfg['distinct_values']}, found {n_import}"
        )
    n_multiple = (fact_lng_contract["importer_raw"] == "Multiple").sum()
    if n_multiple != import_cfg["Multiple"]:
        raise AssertionError(
            f"3.1 distribution assertion failed for import_country=Multiple: "
            f"expected {import_cfg['Multiple']}, found {n_multiple}"
        )


def resolve_country_columns(
    df: pd.DataFrame,
    source_system: str,
    applied_crosswalk: pd.DataFrame,
    raw_columns: dict[str, str],
) -> pd.DataFrame:
    """Resolve one or more raw country columns to ``*_iso2`` columns through
    the applied alias crosswalk for ``source_system``. An unresolved raw
    value raises and names the offender -- build plan 4.8 check 2 already
    proved every raw string resolves, so a miss here means this reader's
    ``source_system`` tag does not match what session 2 used, not a real gap.

    ``raw_columns`` maps {raw_column_name: output_iso2_column_name}.
    """
    xwalk = applied_crosswalk[applied_crosswalk["source_system"] == source_system]
    lookup = xwalk.set_index("raw_value")["country_iso2"]

    out = df.copy()
    for raw_col, iso2_col in raw_columns.items():
        resolved = out[raw_col].map(lookup)
        unresolved = out.loc[resolved.isnull(), raw_col].unique()
        if len(unresolved) > 0:
            raise ValueError(
                f"resolve_country_columns: {source_system}.{raw_col} has raw values with "
                f"no crosswalk entry: {sorted(unresolved.tolist())[:10]}"
            )
        out[iso2_col] = resolved
    return out
