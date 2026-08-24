"""
workbook_reader.py
-------------------
Session 1, build plan 4.1. One reader per LNG workbook. Every sheet is read;
the header row is located, never assumed, and an unexpected or missing
column raises and names the file and sheet. The country-like columns on the
data sheet are extracted as raw strings for ``raw_country_strings.parquet``;
nothing here resolves them, that is section 4.4.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import yaml

_CONFIG_ROOT = Path(__file__).resolve().parents[3] / "config"
with (_CONFIG_ROOT / "session_01_constants.yaml").open(encoding="utf-8") as _f:
    _CONSTANTS = yaml.safe_load(_f)

MAX_HEADER_SEARCH_ROWS = _CONSTANTS["header_search"]["max_rows_scanned"]


@dataclass(frozen=True)
class SheetSchema:
    sheet_name: str
    required_columns: tuple[str, ...]
    country_columns: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookSpec:
    source_system: str
    filename: str
    data_sheet: SheetSchema


WORKBOOK_SPECS: tuple[WorkbookSpec, ...] = (
    WorkbookSpec(
        source_system="workbook_liquefaction",
        filename="202608_LNG_liquefaction_projects.xlsx",
        data_sheet=SheetSchema(
            sheet_name="Global Liquefaction Database",
            required_columns=(
                "Region",
                "Country",
                "ISO 2-letter code",
                "Project",
                "Trains",
                "Company",
                "Start date",
                "Start Year",
                "MTPA",
                "bcf/d",
                "bcma",
                "Status",
                "Type",
            ),
            country_columns=("Country", "ISO 2-letter code"),
        ),
    ),
    WorkbookSpec(
        source_system="workbook_regas",
        filename="202608_LNG_regas_projects.xlsx",
        data_sheet=SheetSchema(
            sheet_name="Global Regas Database",
            required_columns=(
                "Region",
                "Country",
                "ISO 2-letter code",
                "Project",
                "Trains",
                "Company",
                "Start date",
                "Start Year",
                "MTPA",
                "bcf/d",
                "bcma",
                "Status",
                "Type",
            ),
            country_columns=("Country", "ISO 2-letter code"),
        ),
    ),
    WorkbookSpec(
        source_system="workbook_contracts",
        filename="202608_LNG_contracts_database.xlsx",
        data_sheet=SheetSchema(
            sheet_name="Global LNG Contract Database",
            required_columns=(
                "Export country",
                "Supply Area",
                "Import country",
                "Demand Area",
                "Loading Point",
                "Liquefaction project",
                "Project partners",
                "Liquefaction project capacity",
                "FID status",
                "Seller",
                "Buyer",
                "bcm",
                "MTPA",
                "Contract Start ",
            ),
            country_columns=("Export country", "Import country"),
        ),
    ),
)


def _find_header_row(ws, required_columns: tuple[str, ...]) -> tuple[int, list]:
    required = set(required_columns)
    for row_idx in range(1, MAX_HEADER_SEARCH_ROWS + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
        if row is None:
            break
        present = {v for v in row if v is not None}
        if required.issubset(present):
            return row_idx, list(row)
    raise ValueError(
        f"header row not found in sheet '{ws.title}': expected columns "
        f"{sorted(required)} were not all present within the first "
        f"{MAX_HEADER_SEARCH_ROWS} rows"
    )


def _last_nonblank_row(ws, header_row: int) -> int:
    last = header_row
    for i, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        if any(v is not None for v in row):
            last = i
    return last


def _read_non_data_sheet(ws) -> dict:
    row_count = 0
    col_count = 0
    for row in ws.iter_rows(values_only=True):
        non_null = [v for v in row if v is not None]
        if non_null:
            row_count += 1
            col_count = max(col_count, len(non_null))
    return {"row_count": row_count, "column_names": [], "header_row": None}


def read_workbook(path: Path, spec: WorkbookSpec) -> dict:
    """Read every sheet of one workbook.

    Returns a dict with:
      - "sheets": {sheet_name: {"row_count", "column_names", "header_row"}}
      - "country_records": list of {"source_system", "column_name", "raw_value",
        "sheet_name"} for every row in the data sheet's country-like columns.

    Raises ``ValueError`` naming the file and sheet if the expected header
    columns are absent, or if a country column has an empty cell inside the
    data range (fail loudly, no silent drop).
    """
    if not path.is_file():
        raise FileNotFoundError(f"workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets_info: dict[str, dict] = {}
    country_records: list[dict] = []
    try:
        if spec.data_sheet.sheet_name not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: expected data sheet '{spec.data_sheet.sheet_name}' "
                f"not found; sheets present are {wb.sheetnames}"
            )
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name != spec.data_sheet.sheet_name:
                sheets_info[sheet_name] = _read_non_data_sheet(ws)
                continue

            header_row, header_values = _find_header_row(ws, spec.data_sheet.required_columns)
            last_row = _last_nonblank_row(ws, header_row)
            col_index = {name: idx for idx, name in enumerate(header_values) if name is not None}
            sheets_info[sheet_name] = {
                "row_count": last_row - header_row,
                "column_names": [c for c in header_values if c is not None],
                "header_row": header_row,
            }

            for country_col in spec.data_sheet.country_columns:
                idx = col_index[country_col]
                for r, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1, max_row=last_row, values_only=True),
                    start=header_row + 1,
                ):
                    value = row[idx]
                    if value is None:
                        raise ValueError(
                            f"{path.name} sheet '{sheet_name}' row {r}: column "
                            f"'{country_col}' is empty inside the data range "
                            f"({header_row + 1}-{last_row})"
                        )
                    country_records.append(
                        {
                            "source_system": spec.source_system,
                            "column_name": country_col,
                            "raw_value": str(value),
                            "sheet_name": sheet_name,
                        }
                    )
    finally:
        wb.close()

    return {"sheets": sheets_info, "country_records": country_records}


IEA_GTF_SOURCE_SYSTEM = "iea_gtf"
IEA_GTF_DATA_SHEET = "GTF_data"
IEA_GTF_REQUIRED_COLUMNS = ("Borderpoint", "Exit", "Entry", "MAXFLOW (Mm3/h)")
IEA_GTF_COUNTRY_COLUMNS = ("Exit", "Entry")


def read_iea_gtf(path: Path) -> dict:
    """Read the IEA GTF export.

    Only the ``GTF_data`` sheet carries country-like values that feed the
    model (Exit and Entry). The ``NOTES`` sheet is annotation text keyed by
    country combinations such as ``Austria / Germany`` or the literal values
    ``General`` and ``None``; those are not atomic country strings and carry
    no fact this session resolves, so they are recorded in the sheet manifest
    but excluded from ``country_records``.
    """
    if not path.is_file():
        raise FileNotFoundError(f"IEA GTF export not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets_info: dict[str, dict] = {}
    country_records: list[dict] = []
    try:
        if IEA_GTF_DATA_SHEET not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: expected data sheet '{IEA_GTF_DATA_SHEET}' not "
                f"found; sheets present are {wb.sheetnames}"
            )
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name != IEA_GTF_DATA_SHEET:
                sheets_info[sheet_name] = _read_non_data_sheet(ws)
                continue

            header_row, header_values = _find_header_row(ws, IEA_GTF_REQUIRED_COLUMNS)
            last_row = _last_nonblank_row(ws, header_row)
            col_index = {name: idx for idx, name in enumerate(header_values) if name is not None}
            sheets_info[sheet_name] = {
                "row_count": last_row - header_row,
                "column_names": [c for c in header_values if c is not None],
                "header_row": header_row,
            }

            for country_col in IEA_GTF_COUNTRY_COLUMNS:
                idx = col_index[country_col]
                for r, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1, max_row=last_row, values_only=True),
                    start=header_row + 1,
                ):
                    value = row[idx]
                    if value is None:
                        raise ValueError(
                            f"{path.name} sheet '{sheet_name}' row {r}: column "
                            f"'{country_col}' is empty inside the data range "
                            f"({header_row + 1}-{last_row})"
                        )
                    country_records.append(
                        {
                            "source_system": IEA_GTF_SOURCE_SYSTEM,
                            "column_name": country_col,
                            "raw_value": str(value),
                            "sheet_name": sheet_name,
                        }
                    )
    finally:
        wb.close()

    return {"sheets": sheets_info, "country_records": country_records}


def read_project_capacity(
    path: Path, spec: WorkbookSpec, capacity_column: str = "MTPA"
) -> list[dict]:
    """Session 2, build plan 4.3/2.6. One row per project on the liquefaction
    or regas data sheet: raw ``Country``, ``ISO 2-letter code`` and the
    requested capacity column, alongside the project name. Only used for the
    session 2 node/capacity aggregation check, so it reads a single named
    capacity column rather than the whole row -- if a session later needs
    more columns, extend the read here rather than re-deriving from
    ``country_records``, which drops everything except the country strings.
    """
    if not path.is_file():
        raise FileNotFoundError(f"workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    try:
        ws = wb[spec.data_sheet.sheet_name]
        required = (*spec.data_sheet.required_columns, capacity_column)
        header_row, header_values = _find_header_row(ws, required)
        last_row = _last_nonblank_row(ws, header_row)
        col_index = {name: idx for idx, name in enumerate(header_values) if name is not None}
        country_idx = col_index["Country"]
        iso2_idx = col_index["ISO 2-letter code"]
        project_idx = col_index["Project"]
        capacity_idx = col_index[capacity_column]

        for r, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=last_row, values_only=True),
            start=header_row + 1,
        ):
            capacity = row[capacity_idx]
            if capacity is None:
                raise ValueError(
                    f"{path.name} row {r}: column '{capacity_column}' is empty inside "
                    f"the data range ({header_row + 1}-{last_row})"
                )
            rows.append(
                {
                    "source_system": spec.source_system,
                    "project": row[project_idx],
                    "country_raw": row[country_idx],
                    "iso2_raw": row[iso2_idx],
                    "capacity": float(capacity),
                }
            )
    finally:
        wb.close()
    return rows


def read_iea_gtf_border_pairs(path: Path) -> list[dict]:
    """Session 2, build plan 4.3/2.5. One row per GTF border point: the raw
    ``Exit`` and ``Entry`` country strings for that row, paired (unlike
    ``country_records``, which flattens Exit and Entry into separate rows
    and loses the pairing). Used only as evidence for
    ``crosswalks/adjacency_override.csv``: a border point whose Exit/Entry
    countries are not geometrically adjacent is a named pipeline connection
    across a gap the boundary geometry does not show (e.g. a subsea line),
    not proof by itself -- every row here still needs a human-reviewed note.
    """
    if not path.is_file():
        raise FileNotFoundError(f"IEA GTF export not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    try:
        ws = wb[IEA_GTF_DATA_SHEET]
        header_row, header_values = _find_header_row(ws, IEA_GTF_REQUIRED_COLUMNS)
        last_row = _last_nonblank_row(ws, header_row)
        col_index = {name: idx for idx, name in enumerate(header_values) if name is not None}
        exit_idx = col_index["Exit"]
        entry_idx = col_index["Entry"]
        border_idx = col_index["Borderpoint"]

        for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, values_only=True):
            rows.append(
                {
                    "borderpoint": row[border_idx],
                    "exit_raw": row[exit_idx],
                    "entry_raw": row[entry_idx],
                }
            )
    finally:
        wb.close()
    return rows


def read_all_workbooks(workbook_root: Path) -> dict:
    """Read all three pinned LNG workbooks under ``workbook_root``.

    Returns {"per_workbook": {filename: read_workbook(...) result}}.
    """
    per_workbook = {}
    for spec in WORKBOOK_SPECS:
        path = workbook_root / spec.filename
        per_workbook[spec.filename] = read_workbook(path, spec)
    return per_workbook
