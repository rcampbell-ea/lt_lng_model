"""
gtf_flows.py
------------
Session 3, build plan 3.3/3.4/5.3/5.4. ``fact_pipe_flow_hist`` from the IEA
Gas Trade Flows export: physical, directional border-point flows, aggregated
to (``origin_iso2``, ``destination_iso2``, ``year``) in bcm. This is the
European pipeline baseline (plan 5.4) -- Exit is the physical origin of the
flow at that border point, Entry the physical destination, never a
beneficial-origin attribution.

Traps from plan 3.4 handled here:
- Virtual border points (VIP Iberico, Virtualys, VIP BENE, VIP-TH,
  Bras-Petange) are not separate corridors; aggregating straight to country
  pair, as this module does, absorbs the mid-series reallocation rather than
  treating a border point as a stable series in its own right.
- ``Liquefied Natural Gas`` and ``Not Elsewhere Specified`` resolve to the
  XL/XN pseudo codes through the applied crosswalk (session 2), same as
  every other Exit/Entry string; they are never dropped and never enter
  ``dim_country`` as real.
- Reference conditions (15 C, 760 mm Hg) and the monthly Mm3-to-bcm scale
  factor are read from config, never inlined (CLAUDE.md).
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

from lt_lng_flows.ingest.workbook_reader import IEA_GTF_DATA_SHEET, IEA_GTF_REQUIRED_COLUMNS

_MONTH_COL_RE = re.compile(r"^[A-Za-z]{3}-\d{2}$")
_MONTH_ABBR = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}


def _month_col_to_year(col_name: str) -> int:
    abbr, yy = col_name.split("-")
    year = 2000 + int(yy)
    return year


def read_gtf_border_flows(path: Path) -> pd.DataFrame:
    """Long form: one row per (borderpoint, exit_raw, entry_raw, month_col,
    year, value_mm3). ``month_col`` is kept so a caller can distinguish
    within-year months if ever needed; this session only uses ``year``.
    """
    if not path.is_file():
        raise FileNotFoundError(f"IEA GTF export not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if IEA_GTF_DATA_SHEET not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: expected data sheet '{IEA_GTF_DATA_SHEET}' not found; "
                f"sheets present are {wb.sheetnames}"
            )
        ws = wb[IEA_GTF_DATA_SHEET]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = list(header_row)
        missing = [c for c in IEA_GTF_REQUIRED_COLUMNS if c not in header]
        if missing:
            raise ValueError(f"{path.name}: missing required columns {missing}")

        col_index = {name: idx for idx, name in enumerate(header) if name is not None}
        month_cols = [
            name for name in header if isinstance(name, str) and _MONTH_COL_RE.match(name)
        ]

        border_idx = col_index["Borderpoint"]
        exit_idx = col_index["Exit"]
        entry_idx = col_index["Entry"]

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[border_idx] is None:
                continue
            borderpoint = row[border_idx]
            exit_raw = row[exit_idx]
            entry_raw = row[entry_idx]
            for month_col in month_cols:
                value = row[col_index[month_col]]
                if value is None:
                    continue
                if isinstance(value, str):
                    # "N/A" and similar text markers appear where a border
                    # point had no reporting that month (plan 3.4 NOTES
                    # sheet: incomplete collection, discontinued
                    # publication). Not a zero flow and not a parse error;
                    # skip the cell rather than raise or fabricate a value.
                    continue
                records.append(
                    {
                        "borderpoint": borderpoint,
                        "exit_raw": exit_raw,
                        "entry_raw": entry_raw,
                        "month_col": month_col,
                        "year": _month_col_to_year(month_col),
                        "value_mm3": float(value),
                    }
                )
    finally:
        wb.close()
    return pd.DataFrame(records)


def build_fact_pipe_flow_hist(
    gtf_long: pd.DataFrame, applied_crosswalk: pd.DataFrame, mm3_to_bcm: float
) -> pd.DataFrame:
    """Aggregate the long-form monthly border-point flows to
    (``origin_iso2``, ``destination_iso2``, ``year``) in bcm. ``origin_iso2``
    is the resolved Exit code, ``destination_iso2`` the resolved Entry code
    -- physical border flow, plan 5.4's default, not a beneficial-origin
    view. Self pairs (origin == destination) do not occur in the GTF file in
    practice but are not filtered here; if one appeared it would be a data
    quality finding to report, not to silently drop.
    """
    xwalk = applied_crosswalk[applied_crosswalk["source_system"] == "iea_gtf"]
    lookup = xwalk.set_index("raw_value")["country_iso2"]

    df = gtf_long.copy()
    df["origin_iso2"] = df["exit_raw"].map(lookup)
    df["destination_iso2"] = df["entry_raw"].map(lookup)

    unresolved_exit = df.loc[df["origin_iso2"].isnull(), "exit_raw"].unique()
    unresolved_entry = df.loc[df["destination_iso2"].isnull(), "entry_raw"].unique()
    if len(unresolved_exit) or len(unresolved_entry):
        raise ValueError(
            f"build_fact_pipe_flow_hist: unresolved GTF Exit/Entry values -- "
            f"exit={sorted(unresolved_exit.tolist())}, entry={sorted(unresolved_entry.tolist())}"
        )

    grouped = (
        df.groupby(["origin_iso2", "destination_iso2", "year"], as_index=False)["value_mm3"]
        .sum()
        .rename(columns={"value_mm3": "value_mm3_sum"})
    )
    grouped["bcm"] = grouped["value_mm3_sum"] * mm3_to_bcm
    grouped = grouped.drop(columns=["value_mm3_sum"])
    grouped["source"] = "iea_gtf_202606"
    return grouped.sort_values(["origin_iso2", "destination_iso2", "year"]).reset_index(drop=True)
